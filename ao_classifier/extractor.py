"""Extraction d'un échantillon de contenu textuel par type de fichier."""

import logging
import zipfile
from pathlib import Path

from ao_classifier.config import CONTENT_SAMPLE_MAX_CHARS
from ao_classifier.scanner import FileInfo

logger = logging.getLogger(__name__)


def extract_sample(info: FileInfo) -> str:
    """Retourne un échantillon de contenu pour orienter le classement LLM."""
    if info.path.name.startswith("._"):
        return ""
    ext = info.extension
    try:
        if ext == ".pdf":
            return _extract_pdf(info.path)
        elif ext == ".docx":
            return _extract_docx(info.path)
        elif ext in {".xlsx", ".xls"}:
            return _extract_xlsx(info.path)
        elif ext == ".zip":
            return _extract_zip(info.path)
        elif ext == ".7z":
            return _extract_7z(info.path)
        elif ext in {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"}:
            return _extract_image(info.path)
        else:
            return ""
    except Exception as e:
        logger.warning("Extraction échouée pour %s : %s", info.path.name, e)
        return f"[Extraction impossible : {e}]"


def _extract_pdf(path: Path) -> str:
    try:
        import pypdf
        reader = pypdf.PdfReader(str(path))
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
            if len(text) >= CONTENT_SAMPLE_MAX_CHARS:
                break
        text = text.strip()
        if text:
            return text[:CONTENT_SAMPLE_MAX_CHARS]
    except Exception as e:
        logger.debug("pypdf échoué pour %s : %s — tentative pdfplumber", path.name, e)

    # Fallback pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            text = ""
            for page in pdf.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted
                if len(text) >= CONTENT_SAMPLE_MAX_CHARS:
                    break
        text = text.strip()
        if text:
            return text[:CONTENT_SAMPLE_MAX_CHARS]
    except Exception as e:
        logger.debug("pdfplumber échoué pour %s : %s", path.name, e)

    # Fallback Tesseract OCR (PDF scanné)
    return _ocr_pdf(path)


def _extract_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return text[:CONTENT_SAMPLE_MAX_CHARS]


def _extract_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    lines: list[str] = [f"Feuilles : {', '.join(wb.sheetnames)}"]
    for sheet_name in wb.sheetnames[:3]:
        ws = wb[sheet_name]
        rows_extracted = 0
        for row in ws.iter_rows(max_row=5, values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(f"[{sheet_name}] {' | '.join(cells)}")
                rows_extracted += 1
            if rows_extracted >= 4:
                break
    return "\n".join(lines)[:CONTENT_SAMPLE_MAX_CHARS]


def _extract_zip(path: Path) -> str:
    try:
        with zipfile.ZipFile(str(path), "r") as zf:
            names = zf.namelist()
        return "Contenu archive ZIP :\n" + "\n".join(names[:50])
    except zipfile.BadZipFile:
        return "[Archive ZIP corrompue ou invalide]"


def _extract_7z(path: Path) -> str:
    try:
        import py7zr
        with py7zr.SevenZipFile(str(path), mode="r") as z:
            names = z.getnames()
        return "Contenu archive 7z :\n" + "\n".join(names[:50])
    except Exception as e:
        return f"[Archive 7z non lisible : {e}]"


def _ocr_pdf(path: Path) -> str:
    """OCR Tesseract sur un PDF scanné via pdf2image."""
    try:
        from pdf2image import convert_from_path
        import pytesseract

        pages = convert_from_path(str(path), dpi=150, first_page=1, last_page=5)
        text = ""
        for page_img in pages:
            text += pytesseract.image_to_string(page_img, lang="fra+eng") or ""
            if len(text) >= CONTENT_SAMPLE_MAX_CHARS:
                break
        text = text.strip()
        if text:
            logger.debug("OCR Tesseract réussi pour %s (%d chars)", path.name, len(text))
            return text[:CONTENT_SAMPLE_MAX_CHARS]
        return "[PDF scanné — OCR sans résultat]"
    except ImportError:
        return "[PDF scanné — installez pdf2image et pytesseract pour l'OCR]"
    except Exception as e:
        logger.warning("OCR Tesseract échoué pour %s : %s", path.name, e)
        return f"[PDF scanné — OCR échoué : {e}]"


def _extract_image(path: Path) -> str:
    """OCR Tesseract sur une image (PNG, JPG, TIFF…)."""
    try:
        from PIL import Image
        import pytesseract

        img = Image.open(str(path))
        text = pytesseract.image_to_string(img, lang="fra+eng") or ""
        text = text.strip()
        if text:
            return text[:CONTENT_SAMPLE_MAX_CHARS]
        return "[Image — aucun texte détecté par OCR]"
    except ImportError:
        return "[Image — installez Pillow et pytesseract pour l'OCR]"
    except Exception as e:
        logger.warning("OCR image échoué pour %s : %s", path.name, e)
        return f"[Image — OCR échoué : {e}]"


def enrich_files(files: list[FileInfo]) -> list[FileInfo]:
    """Enrichit chaque FileInfo avec un échantillon de contenu."""
    for info in files:
        if info.is_duplicate:
            info.content_sample = "[DOUBLON — contenu identique à un autre fichier]"
            continue
        info.content_sample = extract_sample(info)
        logger.debug(
            "Extrait %d chars depuis %s",
            len(info.content_sample),
            info.path.name,
        )
    return files
